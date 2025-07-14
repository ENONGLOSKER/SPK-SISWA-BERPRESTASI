from django.shortcuts import render, redirect, get_object_or_404
from .models import Siswa, Kriteria, Penilaian, HasilWP, LogAktivitas
from .forms import SiswaForm, KriteriaForm, PenilaianForm
from django.contrib import messages
from django.utils import timezone
from django.db.models import Max
from django.http import JsonResponse
import math
from collections import defaultdict
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models import Count

# Create your views here.

import openpyxl
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from django.urls import reverse
from django.conf import settings

@login_required
def import_siswa(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        # Simpan file sementara
        fs = FileSystemStorage(location=settings.MEDIA_ROOT)
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

            imported = 0
            for row in rows:
                # Asumsi kolom: Nama, NIS, Jenis Kelamin, Tanggal Lahir, Kelas, Alamat
                nama = row[0]
                nis = row[1]
                jenis_kelamin = row[2]
                # Contoh format tgl lahir: "2005-08-17" atau "17/08/2005"
                tanggal_lahir = row[3] if len(row) > 3 else None
                kelas = row[4] if len(row) > 4 else ""
                alamat = row[5] if len(row) > 5 else ""

                if not nama or not nis:
                    continue  # skip baris kosong

                # Cek jika NIS sudah ada
                if Siswa.objects.filter(nis=nis).exists():
                    continue

                # Konversi tanggal lahir jika perlu
                tgl_lahir = None
                if tanggal_lahir:
                    if isinstance(tanggal_lahir, str):
                        # Format string, coba parse
                        from datetime import datetime
                        try:
                            tgl_lahir = datetime.strptime(tanggal_lahir, "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                tgl_lahir = datetime.strptime(tanggal_lahir, "%d/%m/%Y").date()
                            except ValueError:
                                tgl_lahir = None
                    elif hasattr(tanggal_lahir, 'date'):
                        # Excel date
                        tgl_lahir = tanggal_lahir.date() if hasattr(tanggal_lahir, 'date') else tanggal_lahir
                    else:
                        tgl_lahir = tanggal_lahir

                Siswa.objects.create(
                    nama=nama,
                    nis=nis,
                    jenis_kelamin=jenis_kelamin,
                    tanggal_lahir=tgl_lahir,
                    kelas=kelas,
                    alamat=alamat
                )
                imported += 1

            messages.success(request, f'Berhasil mengimpor {imported} data siswa.')
        except Exception as e:
            messages.error(request, f'Gagal mengimpor data: {str(e)}')
        finally:
            fs.delete(filename)
        return redirect('siswa_list')
    return render(request, 'spk/siswa/import_siswa.html', {'title': 'Import Data Siswa'})

from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
import openpyxl
from .models import Siswa, Kriteria, Penilaian
from django.db import transaction

@login_required
def import_penilaian(request):
    """
    Import data penilaian siswa dari file Excel.
    Format Excel:
        Kolom A: NIS
        Kolom B: Nama Siswa (opsional, hanya untuk referensi)
        Kolom C dst: Nilai kriteria (header baris 1 = kode kriteria, baris 2 dst = nilai)
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        fs = FileSystemStorage(location='tmp/')
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Ambil header kode kriteria dari baris pertama
            headers = [cell.value for cell in ws[1]]
            if len(headers) < 3:
                messages.error(request, "Format file tidak valid. Minimal harus ada NIS, Nama, dan satu kriteria.")
                return redirect('import_penilaian')

            # Kriteria dimulai dari kolom ke-3
            kriteria_kode_list = headers[2:]
            kriteria_map = {k.kode: k for k in Kriteria.objects.filter(kode__in=kriteria_kode_list)}
            missing_kriteria = set(kriteria_kode_list) - set(kriteria_map.keys())
            if missing_kriteria:
                messages.error(request, f"Kriteria berikut tidak ditemukan di database: {', '.join(missing_kriteria)}")
                return redirect('import_penilaian')

            imported = 0
            updated = 0
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nis = str(row[0]).strip() if row[0] else None
                    if not nis:
                        continue
                    try:
                        siswa = Siswa.objects.get(nis=nis)
                    except Siswa.DoesNotExist:
                        continue  # skip siswa yang tidak ada

                    for idx, kode in enumerate(kriteria_kode_list):
                        nilai = row[2 + idx]
                        # Pastikan nilai bertipe float (bukan decimal/dll)
                        if nilai is None or (isinstance(nilai, str) and nilai.strip() == ''):
                            continue
                        try:
                            nilai_float = float(nilai)
                        except Exception:
                            continue
                        kriteria = kriteria_map[kode]
                        penilaian_obj, created = Penilaian.objects.update_or_create(
                            siswa=siswa,
                            kriteria=kriteria,
                            defaults={'nilai': nilai_float}
                        )
                        if created:
                            imported += 1
                        else:
                            updated += 1

            messages.success(request, f'Import selesai. {imported} penilaian baru ditambahkan, {updated} penilaian diupdate.')
        except Exception as e:
            messages.error(request, f'Gagal mengimpor data penilaian: {str(e)}')
        finally:
            fs.delete(filename)
        return redirect('penilaian_list')
    return render(request, 'spk/penilaian/import_penilaian.html', {'title': 'Import Data Penilaian'})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            # Catat log aktivitas login
            LogAktivitas.objects.create(
                user=user,
                aksi='login',
                objek='User',
                object_id=user.id,
                keterangan='Login berhasil'
            )
            messages.success(request, "Login berhasil!")
            return redirect('dashboard')
        else:
            messages.error(request, "Username atau password salah.")
    else:
        form = AuthenticationForm()
    return render(request, 'spk/auth/login.html', {'form': form, 'title': 'Login'})

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "Anda telah logout.")
    return redirect('login_view')

@login_required
def dashboard(request):
    total_penilaian = Penilaian.objects.count()
    total_siswa = Siswa.objects.count()
    total_kriteria = Kriteria.objects.count()
    total_bobot = Kriteria.objects.aggregate(total=Sum('bobot'))['total'] or 0

    # Persentase jenis kelamin siswa
    gender_counts = (
        Siswa.objects.values('jenis_kelamin')
        .annotate(jumlah=Count('id'))
        .order_by()
    )
    gender_labels = ['Laki-laki' if g['jenis_kelamin'] == 'L' else 'Perempuan' for g in gender_counts]
    gender_data = [g['jumlah'] for g in gender_counts]

    recent_logs = LogAktivitas.objects.all().order_by('-timestamp')[:4]

    context = {
        'total_penilaian': total_penilaian,
        'total_siswa': total_siswa,
        'total_kriteria': total_kriteria,
        'total_bobot': total_bobot,
        'kategori_labels': gender_labels,
        'kategori_data': gender_data,
        'recent_logs': recent_logs,
    }

    return render(request, 'dashboard.html', context)

# SISWA -------------------------------------------------------------------
@login_required
def siswa_list(request):
    data_siswa = Siswa.objects.all()
    # Catat log aktivitas
    if request.user.is_authenticated:
        LogAktivitas.objects.create(
            user=request.user,
            aksi='lihat',
            objek='Siswa',
            object_id=None,
            keterangan='Melihat daftar siswa'
        )
    context = {
        'data_siswa': data_siswa,
        'title': 'Data Siswa',
        'form_name': 'Data Siswa',
    }
    return render(request, 'spk/siswa/list.html', context)

@login_required
def siswa_add(request):
    if request.method == 'POST':
        form = SiswaForm(request.POST)
        if form.is_valid():
            siswa = form.save()
            # Catat log aktivitas
            LogAktivitas.objects.create(
                user=request.user,
                aksi='tambah',
                objek='Siswa',
                object_id=siswa.id,
                keterangan=f"Menambah siswa: {siswa.nama}"
            )
            messages.success(request, "Data siswa berhasil ditambah!")
            return redirect('siswa_list')
    else:
        form = SiswaForm()
    context = {
        'form': form,
        'title': 'Tambah Siswa',
        'form_name': 'Tambah Siswa',
    }
    return render(request, 'spk/forms/form.html', context)

@login_required
def siswa_update(request, id):
    siswa = get_object_or_404(Siswa, id=id)
    if request.method == 'POST':
        form = SiswaForm(request.POST, instance=siswa)
        if form.is_valid():
            siswa = form.save()
            # Catat log aktivitas
            LogAktivitas.objects.create(
                user=request.user,
                aksi='edit',
                objek='Siswa',
                object_id=siswa.id,
                keterangan=f"Memperbarui data siswa: {siswa.nama}"
            )
            messages.success(request, "Data siswa berhasil diperbarui!")
            return redirect('siswa_list')
    else:
        form = SiswaForm(instance=siswa)
    context = {
        'form': form,
        'title': 'Edit Siswa',
        'form_name': 'Edit Siswa',
    }
    return render(request, 'spk/forms/form.html', context)

@login_required
def siswa_delete(request, id):
    siswa = get_object_or_404(Siswa, id=id)
    nama = siswa.nama
    siswa_id = siswa.id
    siswa.delete()
    # Catat log aktivitas
    LogAktivitas.objects.create(
        user=request.user,
        aksi='hapus',
        objek='Siswa',
        object_id=siswa_id,
        keterangan=f"Data siswa {nama} dihapus"
    )
    messages.success(request, f"Data siswa {nama} berhasil dihapus!")
    return redirect('siswa_list')

# KRITERIA -------------------------------------------------------------------
@login_required
def kriteria_list(request):
    data_kriteria = Kriteria.objects.all()
    total_bobot = data_kriteria.aggregate(total=Sum('bobot'))['total'] or 0
    benefit_count = data_kriteria.filter(jenis='benefit').count()
    cost_count = data_kriteria.filter(jenis='cost').count()
    # Catat log aktivitas
    if request.user.is_authenticated:
        LogAktivitas.objects.create(
            user=request.user,
            aksi='lihat',
            objek='Kriteria',
            object_id=None,
            keterangan='Melihat daftar kriteria'
        )
    context = {
        'data_kriteria': data_kriteria,
        'total_bobot': total_bobot,
        'benefit_count': benefit_count,
        'cost_count': cost_count,
        'title': 'Data Kriteria',
        'form_name': 'Data Kriteria',
    }
    return render(request, 'spk/kriteria/list.html', context)

@login_required
def kriteria_add(request):
    if request.method == 'POST':
        form = KriteriaForm(request.POST)
        if form.is_valid():
            kriteria = form.save()
            # Catat log aktivitas
            LogAktivitas.objects.create(
                user=request.user,
                aksi='tambah',
                objek='Kriteria',
                object_id=kriteria.id,
                keterangan=f"Menambah kriteria: {kriteria.nama}"
            )
            messages.success(request, "Data kriteria berhasil ditambah!")
            return redirect('kriteria_list')
    else:
        form = KriteriaForm()
    context = {
        'form': form,
        'title': 'Tambah Kriteria',
        'form_name': 'Tambah Kriteria',
    }
    return render(request, 'spk/forms/form.html', context)

@login_required
def kriteria_update(request, id):
    kriteria = get_object_or_404(Kriteria, id=id)
    if request.method == 'POST':
        form = KriteriaForm(request.POST, instance=kriteria)
        if form.is_valid():
            kriteria = form.save()
            # Catat log aktivitas
            LogAktivitas.objects.create(
                user=request.user,
                aksi='edit',
                objek='Kriteria',
                object_id=kriteria.id,
                keterangan=f"Memperbarui data kriteria: {kriteria.nama}"
            )
            messages.success(request, "Data kriteria berhasil diperbarui!")
            return redirect('kriteria_list')
    else:
        form = KriteriaForm(instance=kriteria)
    context = {
        'form': form,
        'title': 'Edit Kriteria',
        'form_name': 'Edit Kriteria',
    }
    return render(request, 'spk/forms/form.html', context)

@login_required
def kriteria_delete(request, id):
    kriteria = get_object_or_404(Kriteria, id=id)
    nama = kriteria.nama
    kriteria_id = kriteria.id
    kriteria.delete()
    # Catat log aktivitas
    LogAktivitas.objects.create(
        user=request.user,
        aksi='hapus',
        objek='Kriteria',
        object_id=kriteria_id,
        keterangan=f"Data kriteria {nama} dihapus"
    )
    messages.success(request, f"Data kriteria {nama} berhasil dihapus!")
    return redirect('kriteria_list')

# PENILAIAN -------------------------------------------------------------------
@login_required
def penilaian_list(request):
    data_penilaian = Penilaian.objects.all()
    # Catat log aktivitas
    if request.user.is_authenticated:
        LogAktivitas.objects.create(
            user=request.user,
            aksi='lihat',
            objek='Penilaian',
            object_id=None,
            keterangan='Melihat daftar penilaian'
        )
    context = {
        'data_penilaian': data_penilaian,
        'title': 'Data Penilaian',
        'form_name': 'Data Penilaian',
    }
    return render(request, 'spk/penilaian/list.html', context)

@login_required
def penilaian_add(request):
    if request.method == 'POST':
        form = PenilaianForm(request.POST)
        if form.is_valid():
            penilaian = form.save()
            # Catat log aktivitas
            LogAktivitas.objects.create(
                user=request.user,
                aksi='tambah',
                objek='Penilaian',
                object_id=penilaian.id,
                keterangan=f"Menambah penilaian untuk siswa: {penilaian.siswa.nama}, kriteria: {penilaian.kriteria.nama}, nilai: {penilaian.nilai}"
            )
            messages.success(request, "Data penilaian berhasil ditambah!")
            return redirect('penilaian_list')
    else:
        form = PenilaianForm()
    context = {
        'form': form,
        'title': 'Tambah Penilaian',
        'form_name': 'Tambah Penilaian',
    }
    return render(request, 'spk/forms/form.html', context)

@login_required
def penilaian_update(request, id):
    penilaian = get_object_or_404(Penilaian, id=id)
    if request.method == 'POST':
        form = PenilaianForm(request.POST, instance=penilaian)
        if form.is_valid():
            penilaian = form.save()
            # Catat log aktivitas
            LogAktivitas.objects.create(
                user=request.user,
                aksi='edit',
                objek='Penilaian',
                object_id=penilaian.id,
                keterangan=f"Memperbarui penilaian untuk siswa: {penilaian.siswa.nama}, kriteria: {penilaian.kriteria.nama}, nilai: {penilaian.nilai}"
            )
            messages.success(request, "Data penilaian berhasil diperbarui!")
            return redirect('penilaian_list')
    else:
        form = PenilaianForm(instance=penilaian)
    context = {
        'form': form,
        'title': 'Edit Penilaian',
        'form_name': 'Edit Penilaian',
    }
    return render(request, 'spk/forms/form.html', context)

@login_required
def penilaian_delete(request, id):
    penilaian = get_object_or_404(Penilaian, id=id)
    penilaian_id = penilaian.id
    siswa_nama = penilaian.siswa.nama
    kriteria_nama = penilaian.kriteria.nama
    nilai = penilaian.nilai
    penilaian.delete()
    # Catat log aktivitas
    LogAktivitas.objects.create(
        user=request.user,
        aksi='hapus',
        objek='Penilaian',
        object_id=penilaian_id,
        keterangan=f"Data penilaian untuk siswa: {siswa_nama}, kriteria: {kriteria_nama}, nilai: {nilai} dihapus"
    )
    messages.success(request, "Data penilaian berhasil dihapus!")
    return redirect('penilaian_list')

# HASIL WP -------------------------------------------------------------------
from django.db.models import Sum

@login_required
def hasil_wp_list(request):
    data_hasil = HasilWP.objects.select_related('siswa').all()
    total_data = data_hasil.count()
    total_vector_s = data_hasil.aggregate(total=Sum('vector_s'))['total'] or 0

    siswa_list = Siswa.objects.prefetch_related('penilaian__kriteria')
    penilaian_data = []
    for siswa in siswa_list:
        penilaian_dict = {
            'siswa': siswa,
            'penilaian': {},
            'jumlah_vector_s': 0,
        }
        for penilaian in siswa.penilaian.all():
            kode_kriteria = penilaian.kriteria.kode.upper()
            penilaian_dict['penilaian'][kode_kriteria] = float(penilaian.nilai)
        hasil = data_hasil.filter(siswa=siswa).first()
        if hasil:
            penilaian_dict['jumlah_vector_s'] = float(hasil.vector_s)
        penilaian_data.append(penilaian_dict)

    # Catat log aktivitas
    if request.user.is_authenticated:
        LogAktivitas.objects.create(
            user=request.user,
            aksi='lihat',
            objek='HasilWP',
            object_id=None,
            keterangan='Melihat daftar hasil WP'
        )

    context = {
        'data_hasil': data_hasil,
        'penilaian_data': penilaian_data,
        'total_data': total_data,
        'total_vector_s': float(total_vector_s),
        'title': 'Hasil WP',
        'form_name': 'Hasil WP',
    }
    return render(request, 'spk/hasil_wp/list.html', context)

@login_required
def hasil_wp_detail(request, id):
    hasil = get_object_or_404(HasilWP, id=id)
    # Catat log aktivitas
    LogAktivitas.objects.create(
        user=request.user,
        aksi='lihat',
        objek='HasilWP',
        object_id=hasil.id,
        keterangan=f"Melihat detail hasil WP untuk siswa: {hasil.siswa.nama}"
    )
    context = {
        'hasil': hasil,
        'title': 'Detail Hasil WP',
        'form_name': 'Detail Hasil WP',
    }
    return render(request, 'spk/hasil_wp/detail.html', context)

@login_required
def hasil_wp_delete(request, id):
    hasil = get_object_or_404(HasilWP, id=id)
    hasil_id = hasil.id
    siswa_nama = hasil.siswa.nama
    hasil.delete()
    # Catat log aktivitas
    LogAktivitas.objects.create(
        user=request.user,
        aksi='hapus',
        objek='HasilWP',
        object_id=hasil_id,
        keterangan=f"Data hasil WP untuk siswa: {siswa_nama} dihapus"
    )
    messages.success(request, "Data hasil WP berhasil dihapus!")
    return redirect('hasil_wp_list')

#  perhitungan wp -------------------------------------------------------------------
@login_required
def perhitungan_wp(request):
    if request.method == 'POST':
        if HasilWP.objects.exists():
            messages.info(request, 'Perhitungan WP sudah dilakukan sebelumnya! Silakan reset jika ingin menghitung ulang.')
            return redirect('dashboard')

        siswa_list = Siswa.objects.all()
        kriteria_list = Kriteria.objects.all()
        penilaian_list = Penilaian.objects.select_related('siswa', 'kriteria').all()

        if not siswa_list.exists():
            messages.error(request, 'Data siswa tidak tersedia.')
            return redirect('dashboard')
        if not kriteria_list.exists():
            messages.error(request, 'Data kriteria tidak tersedia.')
            return redirect('dashboard')
        if not penilaian_list.exists():
            messages.error(request, 'Data penilaian tidak tersedia.')
            return redirect('dashboard')

        # Tahap 1: Siapkan bobot ter-normalisasi
        '''
        #Bobot disesuaikan berdasarkan jenis kriteria (benefit atau cost).
        Benefit = 1, Cost = -1
        Contoh matematis normalisasi bobot dan nilai:
        Misal ada 3 kriteria:
          K1 (benefit) bobot=0.4
          K2 (cost)    bobot=0.3
          K3 (benefit) bobot=0.3
        Total bobot = 0.4 + 0.3 + 0.3 = 1.0
        Bobot ternormalisasi:
          K1 = 0.4 / 1.0 = 0.4
          K2 = -0.3 / 1.0 = -0.3
          K3 = 0.3 / 1.0 = 0.3
        
        '''
        total_bobot = sum(float(k.bobot) for k in kriteria_list)
        if total_bobot <= 0:
            messages.error(request, 'Total bobot harus lebih dari 0.')
            return redirect('dashboard')
        bobot_ternormalisasi = {
            k.id: (float(k.bobot) / total_bobot if k.jenis == 'benefit' else -(float(k.bobot) / total_bobot))
            for k in kriteria_list
        }

        # Tahap 2: Struktur nilai penilaian per siswa

        nilai_siswa = defaultdict(dict)
        for p in penilaian_list:
            nilai_siswa[p.siswa_id][p.kriteria_id] = float(p.nilai)

        # Tahap 3: Hitung vector S untuk setiap siswa
        vector_s_dict = {}
        detail_perhitungan = {}

        '''
        CONTOH MATEMATIS:
        Misal ada 2 siswa (A, B) dan 3 kriteria:
        K1 (benefit, bobot=0.4), K2 (cost, bobot=0.3), K3 (benefit, bobot=0.3)
        Bobot ternormalisasi: K1=0.4, K2=-0.3, K3=0.3
        Nilai siswa:
          Siswa A: K1=80, K2=60, K3=90
          Siswa B: K1=70, K2=80, K3=85
        Maka:
          Vector S A = (80^0.4) * (60^-0.3) * (90^0.3)
                     = pangkat1 * pangkat2 * pangkat3
          Vector S B = (70^0.4) * (80^-0.3) * (85^0.3)
                     = pangkat1 * pangkat2 * pangkat3
        Perhitungan kode di bawah ini melakukan hal yang sama untuk seluruh siswa.
        '''
        for siswa in siswa_list:
            vs = 1.0
            detail = {}
            for kriteria in kriteria_list:
                nilai = nilai_siswa.get(siswa.id, {}).get(kriteria.id, 0.0)
                bobot = bobot_ternormalisasi[kriteria.id]
                # pangkat = nilai^bobot (jika nilai > 0, jika tidak maka 0)
                pangkat = math.pow(float(nilai), float(bobot)) if nilai > 0 else 0.0
                vs *= pangkat
                detail[kriteria.kode] = round(float(pangkat), 4)
            vector_s_dict[siswa.id] = float(vs)
            detail_perhitungan[siswa.id] = detail

        # Tahap 4: Hitung vector V
        '''
        CONTOH MATEMATIS:
        Misal ada 2 siswa (A, B) dan 3 kriteria:
        K1 (benefit, bobot=0.4), K2 (cost, bobot=0.3), K3 (benefit, bobot=0.3)
      
        # tahap 1 normalisasi --------------------
        Bobot ternormalisasi: K1=0.4, K2=-0.3, K3=0.3
        Nilai siswa:
          Siswa A: K1=80, K2=60, K3=90
          Siswa B: K1=70, K2=80, K3=85

        # tahap 2 hitung vector s ------------------
        Maka:
            Vector S A = (80^0.4) * (60^-0.3) * (90^0.3)
                        = 3.0314 * 0.4066 * 2.0801
                        = 2.5627
            Vector S B = (70^0.4) * (80^-0.3) * (85^0.3)
                        = 2.7832 × 0.4807 × 2.0166
                        = 2.6998
            total vector s = 2.5627 + 2.6998 = 5.2625
          
        # tahap 3 hitung vector v ------------------
          Vector V A = Vector S A / total seluruh Vector S
          Vector V B = Vector S B / total seluruh Vector S

          Vector V A = 2.5627 / 5.2625
          Vector V B = 2.6998 / 5.2625
          
        Perhitungan kode di bawah ini melakukan hal yang sama untuk seluruh siswa.
        '''
        total_s = sum(float(v) for v in vector_s_dict.values())
        print(f"total_s:{total_s} ===================")
        hasil_list = []
        for siswa in siswa_list:
            vs = float(vector_s_dict[siswa.id])
            vv = vs / total_s if total_s > 0 else 0.0
            hasil_list.append({
                'siswa': siswa,
                'vector_s': round(vs, 6),
                'vector_v': round(vv, 6),
                'detail': detail_perhitungan[siswa.id]
            })

        # ranking
        hasil_list.sort(key=lambda h: h['vector_v'], reverse=True)
        
        for idx, h in enumerate(hasil_list, start=1):
            HasilWP.objects.create(
                siswa=h['siswa'],
                vector_s=float(h['vector_s']),
                vector_v=float(h['vector_v']),
                ranking=idx,
                detail_perhitungan=h['detail'],
                tanggal_hitung=timezone.now(),
                versi="1.0"
            )

        # Catat log aktivitas
        LogAktivitas.objects.create(
            user=request.user,
            aksi='hitung',
            objek='HasilWP',
            object_id=None,
            keterangan='Melakukan perhitungan WP'
        )

        messages.success(request, 'Perhitungan WP berhasil dilakukan!')
        return redirect('hasil_wp')

@login_required
def reset_perhitungan(request):
    HasilWP.objects.all().delete()
    # Catat log aktivitas
    LogAktivitas.objects.create(
        user=request.user,
        aksi='reset',
        objek='HasilWP',
        object_id=None,
        keterangan='Reset perhitungan WP'
    )
    messages.success(request, 'Perhitungan WP berhasil direset!')
    return redirect('dashboard')
